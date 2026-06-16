"""产量/良率/缺陷报表 Excel 导出 Celery 任务"""
import logging
from datetime import date, datetime, timezone

from celery import shared_task
from sqlalchemy import select

from app.core.db import SessionLocal
from app.crud.report_stats import (
    get_daily_trend,
    get_process_rank,
    get_production_summary,
    get_yield_summary,
)
from app.models.quality import DefectCode, InspectionRecord
from app.tasks._excel_utils import fail_job, load_job, save_excel_and_finish, start_job

logger = logging.getLogger(__name__)


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        return None


@shared_task(name="report.production_excel")
def export_production_excel(job_id: int) -> dict:
    """产量报表 Excel 导出：汇总 + 工序排名 + 日趋势"""
    db = SessionLocal()
    try:
        job = load_job(db, job_id)
        if not job:
            return {"ok": False, "msg": "skip"}

        params = start_job(db, job) or {}
        date_from = _parse_date(params.get("date_from"))
        date_to = _parse_date(params.get("date_to"))

        from openpyxl import Workbook
        wb = Workbook()

        # ── Sheet1: 产量汇总 ──
        ws1 = wb.active
        ws1.title = "产量汇总"
        summary = get_production_summary(db, job.tenant_id, date_from=date_from, date_to=date_to)
        ws1.append(["统计指标", "数值"])
        ws1.append(["良品数", summary["good_qty"]])
        ws1.append(["不良品数", summary["bad_qty"]])
        ws1.append(["总产量", summary["total_qty"]])
        ws1.append(["良率", f"{(summary['yield_rate'] or 0) * 100:.2f}%" if summary["yield_rate"] else "—"])
        ws1.append(["报工次数", summary["report_count"]])
        ws1.append(["统计期间", f"{date_from or '全部'} ~ {date_to or '全部'}"])

        # ── Sheet2: 工序排名 ──
        ws2 = wb.create_sheet("工序产量排名")
        ws2.append(["工序", "良品数", "不良品数", "总产量", "良率"])
        rank = get_process_rank(db, job.tenant_id, date_from=date_from, date_to=date_to, limit=200)
        for item in rank:
            yr = f"{item['yield_rate'] * 100:.2f}%" if item["yield_rate"] else "—"
            ws2.append([item["process_name"], item["good_qty"], item["bad_qty"], item["total_qty"], yr])

        # ── Sheet3: 日趋势 ──
        ws3 = wb.create_sheet("日趋势")
        ws3.append(["日期", "良品数", "不良品数", "总产量", "良率"])
        trend = get_daily_trend(db, job.tenant_id, date_from=date_from, date_to=date_to)
        for item in trend:
            yr = f"{item['yield_rate'] * 100:.2f}%" if item["yield_rate"] else "—"
            ws3.append([item["date"], item["good_qty"], item["bad_qty"], item["total_qty"], yr])

        period = f"{date_from or 'all'}_to_{date_to or 'all'}"
        filename = f"production_report_{period}.xlsx"
        return save_excel_and_finish(db, job, wb, filename)

    except Exception as e:
        logger.exception("export_production_excel failed: %s", e)
        return fail_job(db, job_id, e)
    finally:
        db.close()


@shared_task(name="report.yield_excel")
def export_yield_excel(job_id: int) -> dict:
    """良率报表 Excel 导出：良率汇总 + 缺陷 Pareto"""
    db = SessionLocal()
    try:
        job = load_job(db, job_id)
        if not job:
            return {"ok": False, "msg": "skip"}

        params = start_job(db, job) or {}
        date_from = _parse_date(params.get("date_from"))
        date_to = _parse_date(params.get("date_to"))

        from openpyxl import Workbook
        wb = Workbook()

        # ── Sheet1: 良率汇总 ──
        ws1 = wb.active
        ws1.title = "良率汇总"
        yd = get_yield_summary(db, job.tenant_id, date_from=date_from, date_to=date_to)
        ws1.append(["统计指标", "数值"])
        ws1.append(["良品数", yd["good_qty"]])
        ws1.append(["不良品数", yd["bad_qty"]])
        ws1.append(["总产量", yd["total_qty"]])
        ws1.append(["良率", f"{(yd['yield_rate'] or 0) * 100:.2f}%" if yd["yield_rate"] else "—"])

        # ── Sheet2: 缺陷 Pareto ──
        ws2 = wb.create_sheet("缺陷Pareto")
        ws2.append(["缺陷代码", "缺陷名称", "严重程度", "出现次数", "占比%", "累计占比%"])

        from sqlalchemy import func
        from datetime import timedelta
        today = date.today()
        d_from = date_from or (today - timedelta(days=30))
        d_to = date_to or today
        rows = db.execute(
            select(
                DefectCode.code, DefectCode.name, DefectCode.severity,
                func.count(InspectionRecord.id).label("cnt"),
            )
            .select_from(InspectionRecord)
            .join(DefectCode, DefectCode.id == InspectionRecord.defect_code_id)
            .where(
                InspectionRecord.tenant_id == job.tenant_id,
                InspectionRecord.result == "fail",
                InspectionRecord.defect_code_id.isnot(None),
                func.date(InspectionRecord.created_at) >= d_from,
                func.date(InspectionRecord.created_at) <= d_to,
            )
            .group_by(DefectCode.id, DefectCode.code, DefectCode.name, DefectCode.severity)
            .order_by(func.count(InspectionRecord.id).desc())
        ).all()
        total = sum(int(r.cnt) for r in rows)
        cumulative = 0
        for r in rows:
            cnt = int(r.cnt)
            cumulative += cnt
            pct = round(cnt / total * 100, 1) if total else 0
            cum_pct = round(cumulative / total * 100, 1) if total else 0
            ws2.append([r.code, r.name, r.severity, cnt, pct, cum_pct])
        ws2.append([])
        ws2.append(["合计", "", "", total, 100.0, 100.0])

        period = f"{date_from or 'all'}_to_{date_to or 'all'}"
        filename = f"yield_report_{period}.xlsx"
        return save_excel_and_finish(db, job, wb, filename)

    except Exception as e:
        logger.exception("export_yield_excel failed: %s", e)
        return fail_job(db, job_id, e)
    finally:
        db.close()
