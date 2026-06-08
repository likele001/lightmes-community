from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.deps import get_current_user, get_db, require_permissions
from app.core.response import ok
from app.crud.product import get_product_by_id
from app.crud.sku import create_sku, get_sku_by_code, get_sku_by_id, list_skus, update_sku
from app.models.product import Product
from app.models.user import User
from app.schemas.sku import SkuCreateIn, SkuUpdateIn
from app.services.code_generator import BizType, resolve_code
from app.services.display_label import sku_option_extra_fields
from app.services.entity_refs import product_ref_dict


router = APIRouter(dependencies=[Depends(require_permissions(["sku.manage"]))])


def _out_base(x) -> dict:
    return {
        "id": x.id,
        "tenant_id": x.tenant_id,
        "product_id": x.product_id,
        "code": x.code,
        "name": x.name,
        "color": x.color,
        "material": x.material,
        "spec": x.spec,
        "remark": x.remark,
        "is_active": x.is_active,
        "created_at": x.created_at,
        "updated_at": x.updated_at,
    }


def _out(x, product: Product | None = None) -> dict:
    data = _out_base(x)
    p = product
    if p is None and getattr(x, "product", None) is not None:
        p = x.product
    data.update(
        sku_option_extra_fields(
            product_name=p.name if p else None,
            product_description=p.description if p else None,
            product_code=p.code if p else None,
            product_category=p.category if p else None,
            sku_name=x.name,
            sku_code=x.code,
            sku_color=x.color,
            sku_material=x.material,
            sku_spec=x.spec,
        )
    )
    return data


def _product_map_for_skus(db: Session, tenant_id: int, skus: list) -> dict[int, Product]:
    product_ids = {s.product_id for s in skus}
    if not product_ids:
        return {}
    products = db.scalars(
        select(Product).where(Product.tenant_id == tenant_id, Product.id.in_(product_ids))
    ).all()
    return {p.id: p for p in products}


@router.get("")
def list_api(
    product_id: int | None = Query(default=None),
    keyword: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    include_inactive: bool = Query(default=False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if product_id is not None:
        product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=product_id)
        if not product:
            raise HTTPException(status_code=400, detail="产品不存在")
    items = list_skus(
        db,
        tenant_id=user.tenant_id,
        product_id=product_id,
        keyword=keyword,
        offset=offset,
        limit=limit,
        include_inactive=include_inactive,
    )
    pmap = _product_map_for_skus(db, user.tenant_id, items)
    return ok({"items": [_out(x, pmap.get(x.product_id)) for x in items]})


@router.post("")
def create_api(payload: SkuCreateIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=payload.product_id)
    if not product:
        raise HTTPException(status_code=400, detail="产品不存在")
    sku_code = resolve_code(
        db,
        tenant_id=user.tenant_id,
        biz_type=BizType.SKU,
        code=payload.code,
        exists=lambda c: get_sku_by_code(db, user.tenant_id, c) is not None,
        duplicate_msg="型号编码已存在",
    )
    item = create_sku(
        db,
        tenant_id=user.tenant_id,
        product_id=payload.product_id,
        code=sku_code,
        name=payload.name,
        color=payload.color,
        material=payload.material,
        spec=payload.spec,
        remark=payload.remark,
        is_active=payload.is_active,
    )
    db.commit()
    db.refresh(item)
    product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=item.product_id)
    return ok(_out(item, product))


@router.get("/{sku_id}")
def get_api(sku_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    item = get_sku_by_id(db, tenant_id=user.tenant_id, sku_id=sku_id)
    if not item:
        raise HTTPException(status_code=404, detail="产品型号不存在")
    product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=item.product_id)
    return ok(_out(item, product))


@router.put("/{sku_id}")
def update_api(sku_id: int, payload: SkuUpdateIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    item = get_sku_by_id(db, tenant_id=user.tenant_id, sku_id=sku_id)
    if not item:
        raise HTTPException(status_code=404, detail="产品型号不存在")
    if payload.product_id is not None:
        product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=payload.product_id)
        if not product:
            raise HTTPException(status_code=400, detail="产品不存在")
    if payload.code is not None:
        exists = get_sku_by_code(db, tenant_id=user.tenant_id, code=payload.code)
        if exists and exists.id != item.id:
            raise HTTPException(status_code=400, detail="型号编码已存在")
    update_sku(
        db,
        item,
        product_id=payload.product_id,
        code=payload.code,
        name=payload.name,
        color=payload.color,
        material=payload.material,
        spec=payload.spec,
        remark=payload.remark,
        is_active=payload.is_active,
    )
    db.commit()
    db.refresh(item)
    product = get_product_by_id(db, tenant_id=user.tenant_id, product_id=item.product_id)
    return ok(_out(item, product))


@router.delete("/{sku_id}")
def delete_api(sku_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    item = get_sku_by_id(db, tenant_id=user.tenant_id, sku_id=sku_id)
    if not item:
        raise HTTPException(status_code=404, detail="产品型号不存在")
    update_sku(db, item, is_active=False)
    db.commit()
    return ok()


# ── Excel 批量导入 ──

from io import BytesIO
from openpyxl import load_workbook

from app.crud.product import get_product_by_code


@router.post("/import-excel")
def import_skus_excel_api(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量导入型号（Excel），列：product_code, code, name, color, material, spec, remark"""
    if not file:
        raise HTTPException(status_code=400, detail="请上传文件")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的 Excel 文件")
    raw = file.file.read()
    try:
        wb = load_workbook(BytesIO(raw), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请上传 .xlsx 格式")

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    rows_iter = iter(ws.iter_rows(values_only=True))

    # 跳过表头行
    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 文件无数据行")

    # 列映射
    col_map = {}
    for col_idx, val in enumerate(header):
        if val is None:
            continue
        key = str(val).strip().lower().replace(" ", "_")
        if key in ("product_code", "产品编码", "产品编号"):
            col_map["product_code"] = col_idx
        elif key in ("code", "型号编码", "编码"):
            col_map["code"] = col_idx
        elif key in ("name", "型号名称", "名称"):
            col_map["name"] = col_idx
        elif key in ("color", "颜色"):
            col_map["color"] = col_idx
        elif key in ("material", "材料", "材质"):
            col_map["material"] = col_idx
        elif key in ("spec", "规格"):
            col_map["spec"] = col_idx
        elif key in ("remark", "备注"):
            col_map["remark"] = col_idx

    if "code" not in col_map or "name" not in col_map:
        raise HTTPException(
            status_code=400,
            detail='缺少必要列。表头至少包含 "型号编码(code)" 和 "型号名称(name)" 列',
        )

    success = 0
    errors: list[dict] = []
    total_rows = 0

    for row_idx, row in enumerate(rows_iter, start=2):
        total_rows += 1
        code_raw = row[col_map["code"]] if col_map["code"] < len(row) and row[col_map["code"]] is not None else ""
        code = str(code_raw).strip()
        if not code:
            errors.append({"row": row_idx, "message": "型号编码为空"})
            continue

        name_raw = row[col_map["name"]] if col_map["name"] < len(row) and row[col_map["name"]] is not None else ""
        name = str(name_raw).strip()
        if not name:
            errors.append({"row": row_idx, "message": f"型号名称不能为空（编码: {code}）"})
            continue

        # 查找产品
        product_id: int | None = None
        if "product_code" in col_map:
            pc_raw = row[col_map["product_code"]] if col_map["product_code"] < len(row) else None
            if pc_raw is not None:
                pc = str(pc_raw).strip()
                if pc:
                    prod = get_product_by_code(db, tenant_id=user.tenant_id, code=pc)
                    if not prod:
                        errors.append({"row": row_idx, "message": f"产品编码 '{pc}' 不存在（型号: {code}）"})
                        continue
                    product_id = prod.id

        if product_id is None:
            errors.append({"row": row_idx, "message": f"产品编码为空或不存在（型号: {code}）"})
            continue

        # 检查编码重复
        existing = get_sku_by_code(db, tenant_id=user.tenant_id, code=code)
        if existing:
            errors.append({"row": row_idx, "message": f"型号编码 '{code}' 已存在"})
            continue

        # 读取可选字段
        color = None
        if "color" in col_map and col_map["color"] < len(row):
            cv = row[col_map["color"]]
            color = str(cv).strip() if cv is not None else None

        material = None
        if "material" in col_map and col_map["material"] < len(row):
            mv = row[col_map["material"]]
            material = str(mv).strip() if mv is not None else None

        spec = None
        if "spec" in col_map and col_map["spec"] < len(row):
            sv = row[col_map["spec"]]
            spec = str(sv).strip() if sv is not None else None

        remark = None
        if "remark" in col_map and col_map["remark"] < len(row):
            rv = row[col_map["remark"]]
            remark = str(rv).strip() if rv is not None else None

        try:
            create_sku(
                db,
                tenant_id=user.tenant_id,
                product_id=product_id,
                code=code,
                name=name,
                color=color,
                material=material,
                spec=spec,
                remark=remark,
                is_active=True,
            )
            success += 1
        except Exception as e:
            errors.append({"row": row_idx, "message": f"创建失败: {str(e)}（型号: {code}）"})

    db.commit()
    return ok({
        "total": total_rows,
        "success": success,
        "errors": errors,
    })
